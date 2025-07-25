from telethon.sync import TelegramClient
import pandas as pd
import time
from telethon.tl.types import ReactionEmoji, ReactionPaid, ReactionCustomEmoji, MessageMediaPhoto
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
from datetime import datetime, timedelta

# Дата по умолчанию — последние 7 дней
default_end = datetime.now()
default_start = default_end - timedelta(days=7)

# Запрос даты начала
start_input = input(f"Введите дату начала (ГГГГ-ММ-ДД) [по умолчанию {default_start.strftime('%Y-%m-%d')}]: ").strip()
start_date = datetime.strptime(start_input, "%Y-%m-%d") if start_input else default_start

# Запрос даты окончания
end_input = input(f"Введите дату окончания (ГГГГ-ММ-ДД) [по умолчанию {default_end.strftime('%Y-%m-%d')}]: ").strip()
end_date = datetime.strptime(end_input, "%Y-%m-%d") if end_input else default_end

# Конфигурация
api_id = "25492203"
api_hash = "0f1f1ab020df21383b4699bea789871a"
# channel_username = "ekaterinatolasova"
raw_input = input("Введите username канала (с или без @): ").strip()
channel_username = raw_input[1:] if raw_input.startswith("@") else raw_input

# start_date = datetime(2025, 1, 1)
# end_date = datetime(2025, 12, 31, 23, 59, 59)

POSITIVE_EMOJIS = {'👍', '❤', '🔥', '😊', '😂', '🥰', '👏', '⚡', '❤‍🔥', '🫡', '🤗', '😍', '👌', '😁', '💯', '🙏', '🤩'}
NEGATIVE_EMOJIS = {'🤡', '👎', '💩', '🥱'}


def process_reactions(reactions):
    """Обрабатывает реакции и возвращает списки положительных, отрицательных, нейтральных и платных реакций."""
    positive = []
    negative = []
    neutral = []
    paid = []
    total = 0

    if reactions:
        for reaction in reactions.results:
            if isinstance(reaction.reaction, ReactionEmoji):
                emoticon = reaction.reaction.emoticon
                if emoticon in POSITIVE_EMOJIS:
                    positive.append(f"{emoticon} {reaction.count}")
                elif emoticon in NEGATIVE_EMOJIS:
                    negative.append(f"{emoticon} {reaction.count}")
                else:
                    neutral.append(f"{emoticon} {reaction.count}")
            elif isinstance(reaction.reaction, ReactionPaid):
                paid.append(f"{reaction.count}")
            elif isinstance(reaction.reaction, ReactionCustomEmoji):
                neutral.append(f"? {reaction.count}")
        total = sum(reaction.count for reaction in reactions.results)

    return positive, negative, neutral, paid, total


def calculate_engagement_rate(views, reactions, comments, forwards, paid_reactions):
    """Вычисляет Engagement Rate (ER%) с учётом платных реакций."""
    if views > 0:
        total_engagement = reactions + comments + forwards + paid_reactions
        return round(total_engagement / views * 100, 2)
    return 0


def save_to_excel(df, filename, channel_username, start_date, end_date):
    """Сохраняет DataFrame в Excel-файл и добавляет строку с информацией."""
    # Перемещаем строки «Итого» и «Среднее» в начало таблицы
    summary_row = df[df['Дата'] == 'Итого']
    average_row = df[df['Дата'] == 'Среднее']
    df = df[~df['Дата'].isin(['Итого', 'Среднее'])]  # Убираем строки «Итого» и «Среднее» из основного DataFrame
    df = pd.concat([summary_row, average_row, df], ignore_index=True)  # Вставляем их в начало

    df.to_excel(filename, index=False)

    # Загружаем книгу и лист
    wb = load_workbook(filename)
    ws = wb.active

    # Добавляем строку с информацией в начало
    info_text = f"Данные из канала {channel_username} (https://t.me/{channel_username}) за период с {start_date.strftime('%Y-%m-%d %H:%M:%S')} по {end_date.strftime('%Y-%m-%d %H:%M:%S')}"
    ws.insert_rows(1)  # Вставляем строку в начало
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))  # Объединяем ячейки
    ws.cell(row=1, column=1, value=info_text)  # Записываем текст

    # Форматируем строку
    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Светло-голубой цвет
    font = Font(bold=True, size=12)  # Жирный шрифт, размер 12
    alignment = Alignment(horizontal="center", vertical="center")  # Выравнивание по центру

    for cell in ws[1]:  # Применяем стили к первой строке
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment

    # Задаём ширину столбцов
    column_widths = {
        'A': 10,  # Ширина столбца A (например, "Тип")
        'B': 20,  # Ширина столбца B (например, "Дата")
        'C': 30,  # Ширина столбца C (например, "Текст")
        'D': 8,  # Ширина столбца D (например, "Длина")
        'E': 15,  # Ширина столбца E (например, "Ссылка")
        'F': 13,  # Ширина столбца F (например, "Просмотры")
        'G': 11,  # Ширина столбца G (например, "Реакции")
        'H': 17,  # Ширина столбца H (например, "Позитивные")
        'I': 11,  # Ширина столбца I (например, "Всего (+)")
        'J': 17,  # Ширина столбца J (например, "Негативные")
        'K': 11,  # Ширина столбца K (например, "Всего (-)")
        'L': 20,  # Ширина столбца L (например, "Неопределённые")
        'M': 11,  # Ширина столбца M (например, "Всего")
        'N': 13,  # Ширина столбца N (например, "Платные")
        'O': 14,  # Ширина столбца O (например, "Комменты")
        'P': 15,  # Ширина столбца P (например, "Пересылки")
        'Q': 7,  # Ширина столбца Q (например, "ER%")
    }

    # Применяем ширину столбцов
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Закрашиваем первую строку (заголовки)
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Серый цвет
    for cell in ws[2]:  # Вторая строка (после добавленной строки с информацией)
        cell.fill = header_fill

    # Закрашиваем итоговые строки
    summary_fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type="solid")  # Зелёный цвет
    average_fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type="solid")  # Зелёный цвет

    # Итоговая строка (теперь она на третьей строке)
    for cell in ws[3]:  # Третья строка
        cell.fill = summary_fill

    # Строка со средними значениями (теперь она на четвёртой строке)
    for cell in ws[4]:  # Четвёртая строка
        cell.fill = average_fill

    # Замораживаем первые четыре строки (информация, заголовки, итог, среднее)
    ws.freeze_panes = 'A5'  # Теперь закрепляем строки с информацией, заголовками, итогом и средним

    # Сохраняем книгу
    wb.save(filename)


def main():
    """Основная функция для сбора и обработки данных."""
    with TelegramClient('session_name', api_id, api_hash) as client:
        messages = []
        albums = {}

        for message in client.iter_messages(channel_username):
            message_date_naive = message.date.replace(tzinfo=None)

            # Фильтр по дате
            if message_date_naive < start_date:
                break
            if message_date_naive > end_date:
                continue

            # Обработка альбомов
            grouped_id = getattr(message, 'grouped_id', None)
            if grouped_id:
                if grouped_id not in albums:
                    albums[grouped_id] = []
                albums[grouped_id].append(message)
            else:
                # Обработка одиночных сообщений
                post_type = 'Опрос' if message.poll else 'Текст' if message.text else 'Фото' if isinstance(
                    message.media, MessageMediaPhoto) else 'Другое'
                message_link = f"https://t.me/{channel_username}/{message.id}" if channel_username else 'Нет ссылки'

                positive, negative, neutral, paid, total_reactions = process_reactions(message.reactions)
                views = message.views if message.views else 0
                comments = message.replies.replies if message.replies else 0
                forwards = message.forwards if message.forwards else 0

                # Подсчёт платных реакций
                paid_reactions = sum(int(p) for p in paid)
                er_percentage = calculate_engagement_rate(views, total_reactions, comments, forwards, paid_reactions)

                messages.append({
                    'Тип': post_type,
                    'Дата': message_date_naive,
                    'Текст': message.text or '',
                    'Длина': len(message.text or ''),
                    'Ссылка': message_link,
                    'Просмотры': views,
                    'Реакции': total_reactions,
                    'Позитивные': ' '.join(positive),
                    'Всего (+)': sum(int(r.split(' ')[1]) for r in positive),
                    'Негативные': ' '.join(negative),
                    'Всего (-)': sum(int(r.split(' ')[1]) for r in negative),
                    'Неопределённые': ' '.join(neutral),
                    'Всего': sum(int(r.split(' ')[1]) for r in neutral),
                    'Платные': ' '.join(paid),
                    'Комменты': comments,
                    'Пересылки': forwards,
                    'ER%': er_percentage
                })

            # Задержка для соблюдения ограничений API
            time.sleep(0.1)

        # Обработка альбомов
        for grouped_id, album_messages in albums.items():
            first_message = album_messages[0]
            views = first_message.views if first_message.views else 0

            # Суммируем реакции из всех сообщений в альбоме
            positive = []
            negative = []
            neutral = []
            paid = []
            total_reactions = 0
            paid_reactions_total = 0  # Общее количество платных реакций в альбоме

            for message in album_messages:
                p, n, neu, p_paid, total = process_reactions(message.reactions)
                positive.extend(p)
                negative.extend(n)
                neutral.extend(neu)
                paid.extend(p_paid)
                total_reactions += total
                paid_reactions_total += sum(int(p) for p in p_paid)  # Суммируем платные реакции

            comments = sum(m.replies.replies if m.replies else 0 for m in album_messages)
            forwards = sum(m.forwards if m.forwards else 0 for m in album_messages)
            er_percentage = calculate_engagement_rate(views, total_reactions, comments, forwards, paid_reactions_total)

            messages.append({
                'Тип': 'Альбом',
                'Дата': first_message.date.replace(tzinfo=None),
                'Текст': ' '.join(m.text or '' for m in album_messages),
                'Длина': sum(len(m.text or '') for m in album_messages),
                'Ссылка': f"https://t.me/{channel_username}/{first_message.id}" if channel_username else 'Нет ссылки',
                'Просмотры': views,
                'Реакции': total_reactions,
                'Позитивные': ' '.join(positive),
                'Всего (+)': sum(int(r.split(' ')[1]) for r in positive),
                'Негативные': ' '.join(negative),
                'Всего (-)': sum(int(r.split(' ')[1]) for r in negative),
                'Неопределённые': ' '.join(neutral),
                'Всего': sum(int(r.split(' ')[1]) for r in neutral),
                'Платные': ' '.join(paid),
                'Комменты': comments,
                'Пересылки': forwards,
                'ER%': er_percentage
            })

        # Создаём DataFrame
        df = pd.DataFrame(messages)
        df = df.sort_values(by='Дата', ascending=False).reset_index(drop=True)

        # Убедимся, что числовые столбцы имеют правильный тип данных
        numeric_columns = ['Просмотры', 'Реакции', 'Всего (+)', 'Всего (-)', 'Всего', 'Платные', 'Комменты',
                           'Пересылки', 'ER%']
        for col in numeric_columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        # Подсчёт итоговых значений
        summary = {
            'Тип': '',  # Пустое значение для столбца «Тип»
            'Дата': 'Итого',
            'Текст': f'Количество постов: {len(df)}',  # Исправлено: добавляем текст
            'Длина': df['Длина'].sum(),
            'Ссылка': '',
            'Просмотры': df['Просмотры'].sum(),
            'Реакции': df['Реакции'].sum(),
            'Позитивные': '',
            'Всего (+)': df['Всего (+)'].sum(),
            'Негативные': '',
            'Всего (-)': df['Всего (-)'].sum(),
            'Неопределённые': '',
            'Всего': df['Всего'].sum(),
            'Платные': df['Платные'].sum(),
            'Комменты': df['Комменты'].sum(),
            'Пересылки': df['Пересылки'].sum(),
            'ER%': ''
        }

        # Подсчёт средних значений
        average = {
            'Тип': '',  # Пустое значение для столбца «Тип»
            'Дата': 'В среднем на пост',
            'Текст': '',
            'Длина': round(df[df['Длина'] > 0]['Длина'].mean(), 2),
            'Ссылка': '',
            'Просмотры': round(df['Просмотры'].mean(), 2),
            'Реакции': round(df['Реакции'].mean(), 2),
            'Позитивные': '',
            'Всего (+)': round(df['Всего (+)'].mean(), 2),
            'Негативные': '',
            'Всего (-)': round(df['Всего (-)'].mean(), 2),
            'Неопределённые': '',
            'Всего': round(df['Всего'].mean(), 2),
            'Платные': round(df['Платные'].mean(), 2),
            'Комменты': round(df['Комменты'].mean(), 2),
            'Пересылки': round(df['Пересылки'].mean(), 2),
            'ER%': round(df['ER%'].mean(), 2)
        }

        # Получаем порядок столбцов из основного DataFrame
        columns_order = df.columns

        # Создаём DataFrame для «Итого» и «Среднее» с сохранением порядка столбцов
        summary_df = pd.DataFrame([summary], columns=columns_order)
        average_df = pd.DataFrame([average], columns=columns_order)

        # Добавляем итоговые и средние значения в начало DataFrame
        df = pd.concat([summary_df, average_df, df], ignore_index=True)

        # Сохранение в Excel
        filename = f"{channel_username}_{start_date.strftime('%Y-%m-%d')}-{end_date.strftime('%Y-%m-%d')}.xlsx"
        save_to_excel(df, filename, channel_username, start_date, end_date)

        # save_to_excel(df, 'lpr.xlsx', channel_username, start_date, end_date)


if __name__ == "__main__":
    main()
